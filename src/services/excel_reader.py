"""Excel reader — read .xlsx files and parse rows into Appointment models."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.models.appointment import Appointment
from src.utils.excel_column_mapper import ColumnMapping, map_columns

logger = logging.getLogger(__name__)


class ExcelReadError(Exception):
    """Raised when the Excel file cannot be read or is invalid."""

    def __init__(self, message: str, found_columns: list[str] | None = None,
                 missing_columns: list[str] | None = None) -> None:
        super().__init__(message)
        self.found_columns = found_columns or []
        self.missing_columns = missing_columns or []


def _parse_cell_as_str(value: Any) -> str:
    """Convert a cell value to a cleaned string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_start_time(value: Any) -> datetime | None:
    """Parse a cell value as a datetime, handling various formats."""
    if value is None:
        return None

    # If already a datetime object (openpyxl returns datetime for date-formatted cells)
    if isinstance(value, datetime):
        return value

    # If it's a number (Excel serial), convert to datetime
    if isinstance(value, (int, float)):
        # Excel serial date: days since 1900-01-01 (with the 1900 leap year bug)
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value)
        except Exception:
            pass

    # Try string parsing
    text = str(value).strip()
    if not text:
        return None

    # Try ISO format first
    formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    # Try fromisoformat (handles ISO 8601)
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass

    return None


def _parse_duration(value: Any) -> int | None:
    """Parse a cell value as duration in minutes (positive integer)."""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        minutes = int(value)
        return minutes if minutes > 0 else None

    text = str(value).strip()
    if not text:
        return None

    try:
        minutes = int(text)
        return minutes if minutes > 0 else None
    except ValueError:
        # Try extracting digits
        import re
        match = re.search(r"\d+", text)
        if match:
            minutes = int(match.group())
            return minutes if minutes > 0 else None
        return None


def read_excel(
    file_path: str | Path,
    default_country_code: str = "+34",
) -> list[Appointment]:
    """Read an Excel file and return a list of Appointment models.

    Args:
        file_path: Path to the .xlsx file.
        default_country_code: Country code for phone normalization.

    Returns:
        List of Appointment objects (both valid and invalid).

    Raises:
        ExcelReadError: If the file cannot be read or required columns are missing.
    """
    path = Path(file_path)

    if not path.exists():
        raise ExcelReadError(f"No se pudo abrir el archivo: {path}")

    if path.suffix.lower() != ".xlsx":
        raise ExcelReadError("El archivo no es un Excel válido (.xlsx)")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"No se pudo abrir el archivo: {path}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or len(rows) < 2:
        raise ExcelReadError("El archivo no contiene filas de datos")

    # First row = headers
    headers = [str(h) if h is not None else "" for h in rows[0]]
    mapping: ColumnMapping = map_columns(headers)

    if mapping.missing_required:
        found_names = [h for h in headers if h.strip()]
        raise ExcelReadError(
            f"Faltan columnas obligatorias: {', '.join(mapping.missing_required)}. "
            f"Columnas encontradas: {', '.join(found_names)}",
            found_columns=mapping.found_columns,
            missing_columns=mapping.missing_required,
        )

    appointments: list[Appointment] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def get_cell(field: str) -> Any:
            col_idx = mapping.mapping.get(field)
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx]

        patient_name = _parse_cell_as_str(get_cell("patient_name"))
        appointment_type = _parse_cell_as_str(get_cell("appointment_type"))
        gabinete = _parse_cell_as_str(get_cell("gabinete"))
        phone_landline = _parse_cell_as_str(get_cell("phone_landline")) or None
        phone_mobile = _parse_cell_as_str(get_cell("phone_mobile")) or None
        start_time = _parse_start_time(get_cell("start_time"))
        duration = _parse_duration(get_cell("duration"))

        # Build appointment — validation happens in the model
        try:
            appointment = Appointment(
                row_number=row_idx - 1,  # 1-indexed excluding header
                start_time=start_time or datetime(1900, 1, 1),
                duration_minutes=duration or 0,
                gabinete=gabinete,
                patient_name=patient_name,
                appointment_type=appointment_type,
                phone_landline=phone_landline,
                phone_mobile=phone_mobile,
                country_code=default_country_code,
            )
            appointments.append(appointment)
        except Exception as exc:
            logger.warning("Row %d failed to parse: %s", row_idx, exc)
            # Create a minimal invalid appointment to report the error
            appointment = Appointment(
                row_number=row_idx - 1,
                start_time=datetime(1900, 1, 1),
                duration_minutes=1,
                patient_name=patient_name or "(desconocido)",
                appointment_type=appointment_type or "(desconocido)",
                phone_landline=phone_landline,
                phone_mobile=phone_mobile,
                country_code=default_country_code,
            )
            appointments.append(appointment)

    return appointments
